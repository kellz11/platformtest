#!/usr/bin/env python3
"""Rebuild the Core seed workbook from the repository graph."""
import json
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

root = Path(__file__).resolve().parents[1]
json_path = root / "indexer" / "inputs" / "core_graph_curated_100.json"
if not json_path.exists():
    subprocess.run(["python", str(root / "indexer" / "export_graph.py")], check=True)

data = json.loads(json_path.read_text(encoding="utf-8"))
wb = Workbook()
ws = wb.active
ws.title = "nodes"
node_columns = ["curation_rank","id","name","type","parent","description","keywords","related","starter_trend_score","verification_status","confidence_score","canonical","source_plan","source_urls","image_query","images","review_notes"]
ws.append(node_columns)
for node in data["nodes"]:
    ws.append([";".join(node.get(col, [])) if isinstance(node.get(col), list) else node.get(col, "") for col in node_columns])

edges = wb.create_sheet("edges")
edge_columns = ["source","target","relationship","weight","notes","status"]
edges.append(edge_columns)
for edge in data["edges"]:
    edges.append([edge.get(col, "") for col in edge_columns])

for sheet in wb.worksheets:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

out = root / "exports" / "core_graph_seed_v5_indexer_ready.xlsx"
out.parent.mkdir(exist_ok=True)
wb.save(out)
print(out)
